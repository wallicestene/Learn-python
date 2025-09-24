# # Automation in  python

# import openpyxl as xl
# from openpyxl.chart import BarChart, Reference
 


# def process_Workbook(filename):
#     workbook = xl.load_workbook(filename) # loading an excel workbbok. this returns a workbook object
#     sheet = workbook["Sheet1"] #specify the name of the sheet
# # cell = sheet["a1"] # how to access a particular ce  ll #(give the coodinates of the cell that is the combination of the collumn and row)
# # cell = sheet.cell(1, 1) # using the cell method of the sheet object
# # print(cell.value) name of the cell
# # print(sheet.max_row) to get all the rows in that shee
#     for row in range(2, sheet.max_row + 1):
#         cell = sheet.cell(row, 3)
#         correct_price = cell.value * 0.9
#         correct_price_Cell = sheet.cell(row, 4)
#         correct_price_Cell.value = correct_price

#     values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

#     chart = BarChart()

#     chart.add_data(values)

#     sheet.add_chart(chart, "e2")

#     workbook.save(filename)

# process_Workbook("transactions.xlsx")


import openpyxl as xl
from openpyxl.chart import PieChart, Reference

def process_workbook(file_name):
    workbook = xl.load_workbook(file_name)
    sheet = workbook["Sheet1"]

    # print(sheet.max_row)
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 4)
        # print(cell.value)
        new_price = cell.value * 0.99    
        # print(newPrice)
        new_price_cell = sheet.cell(row, 5)
        new_price_cell.value = new_price


    values = Reference(sheet,min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)

    chart = PieChart()

    chart.add_data(values)

    sheet.add_chart(chart, "f2")
        
    workbook.save("transactions2.xlsx")



process_workbook("transactions2.xlsx")